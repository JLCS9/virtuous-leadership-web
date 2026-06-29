#!/usr/bin/env python3
"""Extrae el banco de preguntas, support_text y fixtures del Excel oficial
del test de carácter (Havard).

Lee el archivo `.xlsx` de origen y emite TRES JSONs en src/data/:

  1. character-test.json
      - facets: lista ordenada de las 12 facetas (nombre interno + virtud
        a la que pertenecen + role passive/active).
      - facet_denominators: suma de pesos por faceta (5.1, 3, 9, ...).
      - virtues: 6 virtudes con sus 2 facetas y denominador agregado.
      - questions: 68 ítems {code, text:{es,en,fr,ru}, weights:[12]}.

  2. character-support-text.json
      - keys: dict { key: {es,en,fr,ru} } con todas las claves chResult*,
        las etiquetas auxiliares y los 7 HTMLs largos.

  3. character-test-fixtures.json
      - real: filas de Sheet6 (test_results_character) con sus
        answers_values, locale y resultado esperado (precomputado aquí
        con el mismo motor para coherencia algoritmo↔datos).
      - synthetic: casos límite (all_plus_1, all_minus_1, all_zero).

Uso:
    python3 scripts/extract-character-xlsx.py [path/al/Excel]

El argumento opcional sobrescribe la ruta por defecto del Excel. Los
JSONs se escriben siempre en src/data/.

Determinismo: no random, no fechas. Si el Excel no cambia, los JSONs no
cambian — apto para reproducir builds.
"""

import ast
import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = '/Users/joseluiscastedosanchez/Downloads/TRAD Character test 4.xlsx'
OUT_DIR = REPO_ROOT / 'src' / 'data'

# Orden estable de las 12 facetas — coincide con el orden de las columnas
# F-Q en la hoja `questions` y con el orden semántico
# (cada virtud: pasiva, activa).
FACETS = [
    # (key interno,  virtud,  role,     i18n key support_text)
    ('deliberation',     'P', 'passive', 'chResultP1'),
    ('decision',         'P', 'active',  'chResultP2'),
    ('endurance',        'C', 'passive', 'chResultC1'),
    ('audacity',         'C', 'active',  'chResultC2'),
    ('subduing',         'S', 'passive', 'chResultS1'),
    ('energizing',       'S', 'active',  'chResultS2'),
    ('communion',        'J', 'passive', 'chResultJ1'),
    ('common_good',      'J', 'active',  'chResultJ2'),
    ('contemplation',    'M', 'passive', 'chResultM1'),
    ('action',           'M', 'active',  'chResultM2'),
    ('self_knowledge',   'H', 'passive', 'chResultH1'),
    ('service',          'H', 'active',  'chResultH2'),
]

# 6 virtudes: code → (i18n key, HTML key support_text, denominador esperado).
# El denominador se VALIDA contra la suma real de pesos del Excel; si no
# coincide el script aborta — evita drift silencioso si el Excel cambia.
VIRTUES = [
    ('P', 'chResultP', '1-prudence.html',   8.1),
    ('C', 'chResultC', '2-courage.html',    14),
    ('S', 'chResultS', '3-selfmastery.html', 9),
    ('J', 'chResultJ', '4-justice.html',    9),
    ('M', 'chResultM', '5-magnanimity.html', 18),
    ('H', 'chResultH', '6-humility.html',   15),
]

LANGS = [('ru', 'RU'), ('en', 'EN'), ('fr', 'FR'), ('es', 'ES')]


def load_questions_sheet(wb):
    """Lee la hoja 'questions': 68 filas con texto + matriz de pesos."""
    ws = wb['questions']
    questions = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=5).value
        if not code:
            continue
        # Texto por idioma. Columnas A-D = RU, EN, FR, ES (en ese orden
        # según el header).
        text = {
            'ru': ws.cell(row=r, column=1).value or '',
            'en': ws.cell(row=r, column=2).value or '',
            'fr': ws.cell(row=r, column=3).value or '',
            'es': ws.cell(row=r, column=4).value or '',
        }
        # Pesos: columnas F-Q (6..17) en el orden de FACETS.
        weights = []
        for i in range(12):
            v = ws.cell(row=r, column=6 + i).value
            # openpyxl devuelve enteros o floats; aseguramos float.
            weights.append(float(v) if isinstance(v, (int, float)) else 0.0)
        questions.append({
            'code': str(code).strip(),
            'text': text,
            'weights': weights,
        })
    return questions


def load_support_text(wb):
    """Lee la hoja 'support_text': clave + 4 idiomas (EN/RU/FR/ES)."""
    ws = wb['support_text']
    keys = {}
    # Header row 1: key/locale, EN, RU, FR, ES, PT, DE, IT
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if not key:
            continue
        keys[str(key)] = {
            'en': ws.cell(row=r, column=2).value or '',
            'ru': ws.cell(row=r, column=3).value or '',
            'fr': ws.cell(row=r, column=4).value or '',
            'es': ws.cell(row=r, column=5).value or '',
        }
    return keys


def score(answers, questions, facets, virtues):
    """Motor de scoring puro — mismo algoritmo que characterScoring.js.
    Lo replicamos en Python aquí SOLO para precomputar los resultados
    esperados de las fixtures. Si los dos cálculos divergen, los tests
    de characterScoring.js fallarán.
    """
    if len(answers) != len(questions):
        # Truncamos o rellenamos con 0 — los datos de Sheet6 pueden tener
        # menos de 68 respuestas (tests abandonados).
        answers = (list(answers) + [0.0] * len(questions))[:len(questions)]
    facet_raw = [0.0] * 12
    for i, ans in enumerate(answers):
        if not isinstance(ans, (int, float)):
            continue
        weights = questions[i]['weights']
        for f in range(12):
            facet_raw[f] += ans * weights[f]
    # Denominadores por faceta = suma de pesos de esa columna.
    facet_denoms = [sum(q['weights'][f] for q in questions) for f in range(12)]

    def pct(raw, denom):
        if not denom:
            return None
        norm = raw / denom  # ∈ [-1, +1]
        return round(((norm + 1.0) / 2.0) * 100.0, 2)

    out = {}
    for v_code, _, _, _ in virtues:
        # Encuentra las dos facetas de esta virtud (la pasiva = 1ª, la
        # activa = 2ª en orden de FACETS).
        idx_passive = next(i for i, (_, vc, role, _) in enumerate(facets)
                           if vc == v_code and role == 'passive')
        idx_active = next(i for i, (_, vc, role, _) in enumerate(facets)
                          if vc == v_code and role == 'active')
        raw_p = facet_raw[idx_passive]
        raw_a = facet_raw[idx_active]
        denom_p = facet_denoms[idx_passive]
        denom_a = facet_denoms[idx_active]
        out[v_code] = {
            'global': pct(raw_p + raw_a, denom_p + denom_a),
            'passive': pct(raw_p, denom_p),
            'active': pct(raw_a, denom_a),
        }
    return out


def load_fixtures(wb, questions, facets, virtues):
    """Saca las filas reales de Sheet6 con test_results_character poblado,
    parsea cada bloque (formato Flutter Python-literal) y precomputa el
    resultado esperado.
    """
    ws = wb['Sheet6']
    real = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=12).value  # col 12 = test_results_character
        if not cell or not str(cell).strip():
            continue
        try:
            entries = ast.literal_eval(cell)
        except (ValueError, SyntaxError) as e:
            print(f'  [skip row {r}] parse error: {e}', file=sys.stderr)
            continue
        if not isinstance(entries, list):
            continue
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            raw = entry.get('answers_values')
            if not raw:
                continue
            # Parse comma-separated answers.
            try:
                answers = [float(x.strip()) for x in str(raw).split(',') if x.strip()]
            except ValueError:
                continue
            expected = score(answers, questions, facets, virtues)
            real.append({
                'source_row': r,
                'locale': entry.get('locale', 'unknown'),
                'answers_values': answers,
                'expected': expected,
            })

    # Casos sintéticos (edge cases).
    n = len(questions)
    synthetic = {
        'all_plus_1':  {'answers_values': [1.0] * n,
                        'expected': score([1.0] * n, questions, facets, virtues)},
        'all_minus_1': {'answers_values': [-1.0] * n,
                        'expected': score([-1.0] * n, questions, facets, virtues)},
        'all_zero':    {'answers_values': [0.0] * n,
                        'expected': score([0.0] * n, questions, facets, virtues)},
    }
    return real, synthetic


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    print(f'Reading {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    questions = load_questions_sheet(wb)
    print(f'  questions: {len(questions)} (expected 68)')
    assert len(questions) == 68, f'expected 68 questions, got {len(questions)}'

    # Verifica que los denominadores reales coinciden con los del briefing.
    expected_facet_sums = [5.1, 3, 9, 5, 6, 3, 4, 5, 6, 12, 9, 6]
    actual = [sum(q['weights'][f] for q in questions) for f in range(12)]
    for i, (exp, got) in enumerate(zip(expected_facet_sums, actual)):
        if abs(exp - got) > 1e-6:
            raise SystemExit(f'  ✗ facet {FACETS[i][0]}: expected sum {exp}, got {got}')
    print(f'  ✓ 12 facet denominators match briefing')

    # Verifica denominadores por virtud.
    for v_code, _, _, exp_d in VIRTUES:
        idx_p = next(i for i, (_, vc, role, _) in enumerate(FACETS) if vc == v_code and role == 'passive')
        idx_a = next(i for i, (_, vc, role, _) in enumerate(FACETS) if vc == v_code and role == 'active')
        got_d = actual[idx_p] + actual[idx_a]
        if abs(exp_d - got_d) > 1e-6:
            raise SystemExit(f'  ✗ virtue {v_code}: expected denom {exp_d}, got {got_d}')
    print(f'  ✓ 6 virtue denominators match briefing')

    support = load_support_text(wb)
    print(f'  support_text keys: {len(support)}')

    # Compone character-test.json
    test_json = {
        'meta': {
            'source': 'TRAD Character test 4.xlsx',
            'questions_count': len(questions),
            'facets_count': 12,
            'virtues_count': 6,
        },
        'facets': [
            {'key': k, 'virtue': vc, 'role': role, 'label_key': lk,
             'denominator': round(actual[i], 4)}
            for i, (k, vc, role, lk) in enumerate(FACETS)
        ],
        'virtues': [
            {'code': c, 'label_key': lk, 'html_key': hk,
             'denominator': round(d, 4),
             'passive_facet': next(f[0] for f in FACETS if f[1] == c and f[2] == 'passive'),
             'active_facet':  next(f[0] for f in FACETS if f[1] == c and f[2] == 'active')}
            for c, lk, hk, d in VIRTUES
        ],
        'questions': questions,
    }

    # Compone character-support-text.json — separamos las claves cortas
    # de los HTMLs largos para que el componente React decida qué cargar.
    html_keys = [v[2] for v in VIRTUES] + ['text-about-character-test.html']
    support_json = {
        'meta': {
            'source': 'TRAD Character test 4.xlsx → sheet support_text',
            'keys_count': len(support),
        },
        'labels': {k: v for k, v in support.items() if k not in html_keys},
        'html': {k: v for k, v in support.items() if k in html_keys},
    }

    # Fixtures.
    real, synthetic = load_fixtures(wb, questions, FACETS, VIRTUES)
    fixtures_json = {
        'meta': {
            'source': 'TRAD Character test 4.xlsx → sheet Sheet6 (test_results_character)',
            'real_count': len(real),
            'note': '`expected` is precomputed in extract-character-xlsx.py with the same algorithm '
                    'as src/lib/characterScoring.js. Tests assert they match — if the JS engine '
                    'diverges from this Python reference, the unit tests fail.',
        },
        'real': real,
        'synthetic': synthetic,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out1 = OUT_DIR / 'character-test.json'
    out2 = OUT_DIR / 'character-support-text.json'
    out3 = OUT_DIR / 'character-test-fixtures.json'

    out1.write_text(json.dumps(test_json, ensure_ascii=False, indent=2), encoding='utf-8')
    out2.write_text(json.dumps(support_json, ensure_ascii=False, indent=2), encoding='utf-8')
    out3.write_text(json.dumps(fixtures_json, ensure_ascii=False, indent=2), encoding='utf-8')

    print(f'\nWrote:')
    print(f'  {out1}  ({out1.stat().st_size:,} bytes)')
    print(f'  {out2}  ({out2.stat().st_size:,} bytes)')
    print(f'  {out3}  ({out3.stat().st_size:,} bytes — {len(real)} real fixtures, 3 synthetic)')


if __name__ == '__main__':
    main()
